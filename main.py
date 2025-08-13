# 背单词
from tkinter import *
from tkinter import ttk
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import random

addr = r"C:\Users\ruiha\Desktop\1\vocab1.xlsx"# Your directory to the execl
fraction = 10
pool = {}
words_displayed = {}

root = Tk()
root.title("Vocabulary")
root.geometry("600x450")
root.configure(bg="white")
wb = load_workbook(addr)
ws = wb.active

for i, row in enumerate(ws.iter_rows(values_only=True)):
    if row[1]:
        if row[0] == False:
            cell = ws["B" + str(i + 1)]
            word = row[1]
            expl = row[2]
            if cell.fill.start_color.index == "FFFFFF00":
                words_displayed[word] = expl
            else:
                pool[word] = expl
print(pool, len(pool))
chosen = random.sample(list(pool.items()), int(len(pool) / fraction))
for item in chosen:
    words_displayed[item[0]] = item[1]
pages = len(words_displayed)
style = ttk.Style()
style.configure(
    "Custom.TLabel",
    background="lightblue",
    font=("Arial", 20, "bold"),
    bd=2,
    relief="solid"
)


class Window:
    def __init__(self, word_dict):
        self.cur_page = 1
        self.repeated_show = False
        self.word_dict = word_dict
        self.word = ttk.Label(root, text="", style="Custom.TLabel")
        self.page = ttk.Label(root, text=str(self.cur_page) + "/" + str(len(word_dict)), font=("arial", 15))
        self.cur_def = ""
        self.expl = Text(root, height=8, width=60, font=("arial,12"), state=DISABLED)
        self.dropdown = ttk.Combobox(root, values=list(word_dict.keys()), state="readonly")
        self.last = ttk.Button(root, text="Last", command=lambda: self.flip(-1))
        self.show = ttk.Button(root, text="Show Definition", command=self.show)
        self.next = ttk.Button(root, text="next", command=lambda: self.flip(1))
        self.page.place(relx=0.5, rely=0.05, anchor="n")
        self.last.place(relx=0.3, rely=0.55, anchor="center")
        self.next.place(relx=0.7, rely=0.55, anchor="center")
        self.word.place(relx=0.5, rely=0.3, anchor="center")
        self.show.place(relx=0.5, rely=0.55, anchor="center")
        self.expl.pack(padx=0.5, pady=10, side=BOTTOM, fill=X)
        self.dropdown.place(relx=0.7, rely=0.05, anchor="n")
        self.dropdown.set(list(word_dict.keys())[self.cur_page - 1])
        self.dropdown.bind("<<ComboboxSelected>>", lambda event: self.flip(self.dropdown.current() - self.cur_page + 1))

    def flip(self, n):
        self.cur_page += n
        if self.cur_page == -1:
            self.cur_page = len(self.word_dict)
        if self.cur_page > len(self.word_dict):
            self.cur_page = 1
        self.repeated_show = False
        self.render()
        self.expl.config(state=NORMAL)
        self.expl.delete(1.0, END)
        self.expl.config(state=DISABLED)

    def render(self):
        self.word.configure(text=list(self.word_dict.keys())[self.cur_page - 1])
        self.page.configure(text=str(self.cur_page) + "/" + str(len(self.word_dict)))

    def show(self):
        if not self.repeated_show:
            self.expl.config(state=NORMAL)
            self.repeated_show = True
        self.expl.insert(END, list(self.word_dict.values())[self.cur_page - 1])
        self.expl.config(state=DISABLED)


window = Window(words_displayed)
window.render()
print(words_displayed, len(words_displayed))

root.mainloop()
